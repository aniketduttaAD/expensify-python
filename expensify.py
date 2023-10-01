import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkcalendar import DateEntry
import os
import openpyxl
from openpyxl import Workbook
from ttkthemes import ThemedStyle
import matplotlib.pyplot as plt
import numpy as np

category_totals = {}

def handle_focus_in(event):
    if event.widget == transaction_detail_entry and transaction_detail_entry.get() == "Transaction Detail":
        transaction_detail_entry.delete(0, "end")
        transaction_detail_entry.configure(foreground="black")
    elif event.widget == amount_entry and amount_entry.get() == "Amount":
        amount_entry.delete(0, "end")
        amount_entry.configure(foreground="black")

def handle_focus_out(event):
    if event.widget == transaction_detail_entry and not transaction_detail_entry.get():
        transaction_detail_entry.insert(0, "Transaction Detail")
        transaction_detail_entry.configure(foreground="gray")
    elif event.widget == amount_entry and not amount_entry.get():
        amount_entry.insert(0, "Amount")
        amount_entry.configure(foreground="gray")

def create_excel_file(filename):
    if not os.path.exists(filename):
        workbook = Workbook()
        workbook.save(filename)

def calculate_totals():
    excel_filename = "expenses.xlsx"
    category_totals = {}
    transaction_totals = {"Debit": 0, "Credit": 0}

    if os.path.exists(excel_filename):
        workbook = openpyxl.load_workbook(excel_filename)
        worksheet = workbook["Expenses"]

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            _, _, category, amount, transaction_type = row

            if transaction_type in transaction_totals:
                transaction_totals[transaction_type] += amount

                category = category.lower()
                if category in category_totals:
                    category_totals[category] += amount
                else:
                    category_totals[category] = amount

    return transaction_totals, category_totals

def update_totals():
    transaction_totals, category_totals = calculate_totals()
    total_debit = transaction_totals.get("Debit", 0)
    total_credit = transaction_totals.get("Credit", 0)
    balance = total_debit - total_credit
    total_debit_label.config(text=f"Total Debit: {total_debit:.2f}")
    total_credit_label.config(text=f"Total Credit: {total_credit:.2f}")
    balance_label.config(text=f"Balance: {balance:.2f}")

    for i, category in enumerate(categories):
        total = category_totals.get(category.lower(), 0.0)
        category_labels[i].config(text=f"{category.capitalize()} Total: {total:.2f}")

def add_expense():
    transaction_detail = transaction_detail_entry.get()
    category = category_var.get().lower()
    amount = amount_entry.get()
    transaction_type = transaction_type_var.get()
    date = date_entry.get_date()

    if not transaction_detail or not category or not amount:
        messagebox.showerror("Error", "Please fill in all fields.")
        return

    try:
        amount = float(amount)
    except ValueError:
        messagebox.showerror("Error", "Invalid amount. Please enter a valid number.")
        return

    if len(str(amount).split(".")[1]) > 2:
        messagebox.showerror("Error", "Amount should have up to 2 decimal places.")
        return

    expense_entry = f"Expense Date: {date}| Detail: {transaction_detail}| Category: {category}| Amount: {amount:.2f}| Type: {transaction_type}"

    expense_listbox.insert(tk.END, expense_entry)

    if category in category_totals:
        category_totals[category] += amount
    else:
        category_totals[category] = amount

    update_totals()

    transaction_detail_entry.delete(0, tk.END)
    transaction_detail_entry.insert(0, "Transaction Detail")
    transaction_detail_entry.configure(foreground="black")
    category_var.set("")
    amount_entry.delete(0, tk.END)
    amount_entry.insert(0, "Amount")
    amount_entry.configure(foreground="black")
    transaction_type_var.set("")
    date_entry.set_date(None)

    result_label.config(text="Expense added", foreground="white", background="orange")

def submit_expenses():
    expenses = expense_listbox.get(0, tk.END)
    data = []

    for expense in expenses:
        parts = expense.split('| ')
        date = parts[0].split(': ')[1]
        detail = parts[1].split(': ')[1]
        category = parts[2].split(': ')[1]
        amount = float(parts[3].split(': ')[1])
        transaction_type = parts[4].split(': ')[1]

        data.append([date, detail, category, amount, transaction_type])

    excel_filename = "expenses.xlsx"

    if not data:
        return

    if not os.path.exists(excel_filename):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Expenses"
        headers = ["Date", "Detail", "Category", "Amount", "Type"]
        worksheet.append(headers)
    else:
        workbook = openpyxl.load_workbook(excel_filename)
        if "Expenses" not in workbook.sheetnames:
            worksheet = workbook.create_sheet("Expenses")
            headers = ["Date", "Detail", "Category", "Amount", "Type"]
            worksheet.append(headers)
        else:
            worksheet = workbook["Expenses"]

    for row_data in data:
        worksheet.append(row_data)

    workbook.save(excel_filename)

    app.quit()

app = tk.Tk()
app.title("Expense Tracker")
app.geometry("1400x600")
style = ThemedStyle(app)
style.set_theme("scidblue")
app.configure(bg="#333333")
img=tk.Image('photo',file="icon.gif")
app.tk.call('wm','iconphoto', app._w, img)


input_frame = ttk.Frame(app, padding=10, style="Input.TFrame")
input_frame.pack(pady=20)

transaction_detail_entry = ttk.Entry(input_frame, style="Placeholder.TEntry")
transaction_detail_entry.insert(0, "Transaction Detail")
transaction_detail_entry.grid(row=0, column=0, padx=5)
transaction_detail_entry.bind("<FocusIn>", handle_focus_in)
transaction_detail_entry.bind("<FocusOut>", handle_focus_out)

categories = ["Grocery", "Food", "Miscellaneous", "My Expenses", "Dad", "Brother"]
category_var = tk.StringVar()
category_dropdown = ttk.Combobox(input_frame, textvariable=category_var, values=categories, style="Orange.TCombobox")
category_dropdown.set("Category")
category_dropdown.grid(row=0, column=1, padx=5)

amount_entry = ttk.Entry(input_frame, style="Orange.TEntry")
amount_entry.insert(0, "Amount")
amount_entry.grid(row=0, column=2, padx=5)
amount_entry.bind("<FocusIn>", handle_focus_in)
amount_entry.bind("<FocusOut>", handle_focus_out)

transaction_types = ["Debit", "Credit"]
transaction_type_var = tk.StringVar()
transaction_type_dropdown = ttk.Combobox(input_frame, textvariable=transaction_type_var, values=transaction_types)
transaction_type_dropdown.set("Debit")
transaction_type_dropdown.grid(row=0, column=3, padx=5)

date_entry = DateEntry(input_frame, width=12, background="orange", foreground="black", style="Orange.DateEntry")
date_entry.grid(row=0, column=4, padx=5)

add_button = ttk.Button(input_frame, text="Add Expense", command=add_expense, style="OrangeHover.TButton")
add_button.grid(row=0, column=5, padx=5)

listbox_frame = ttk.Frame(app)
listbox_frame.pack(padx=20, fill=tk.BOTH, expand=True)

expense_listbox = tk.Listbox(listbox_frame, selectmode=tk.SINGLE, width=100, height=8)
expense_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

listbox_scrollbar = tk.Scrollbar(listbox_frame, orient=tk.VERTICAL)
listbox_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
expense_listbox.config(yscrollcommand=listbox_scrollbar.set)
listbox_scrollbar.config(command=expense_listbox.yview)

result_label = ttk.Label(app, text="", foreground="green")
result_label.pack(pady=10)

submit_button = ttk.Button(app, text="Submit Expenses", command=submit_expenses, style="OrangeHover.TButton")
submit_button.pack(pady=10)

style = ttk.Style()
style.configure('Input.TFrame', background='#333333')
style.configure('Orange.TCombobox', fieldbackground='#FFF')
style.configure('Orange.TEntry', fieldbackground='#FFF')
style.configure('Orange.DateEntry', fieldbackground='#FFF', arrowcolor='black')
style.configure('Placeholder.TEntry', foreground='black')
style.configure('OrangeHover.TButton', background='#333333', foreground='orange', bordercolor='orange', lightcolor='orange', darkcolor='orange')

totals_frame = ttk.Frame(app, padding=10)
totals_frame.pack(pady=10)

total_debit_label = ttk.Label(totals_frame, text="", foreground="green")
total_debit_label.grid(row=0, column=0, padx=10)

total_credit_label = ttk.Label(totals_frame, text="", foreground="green")
total_credit_label.grid(row=0, column=1, padx=10)

balance_label = ttk.Label(totals_frame, text="", foreground="green")
balance_label.grid(row=0, column=2, padx=10)

category_labels = []
for i, category in enumerate(categories):
    label = ttk.Label(totals_frame, text=f"{category.capitalize()} Total: 0.00", foreground="green")
    label.grid(row=0, column=i + 3, padx=10)
    category_labels.append(label)

def create_category_pie_chart():
    excel_filename = "expenses.xlsx"
    category_totals = {}

    if os.path.exists(excel_filename):
        workbook = openpyxl.load_workbook(excel_filename)
        worksheet = workbook["Expenses"]

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            _, _, category, amount, transaction_type = row

            category = category.lower()
            if category in category_totals:
                if transaction_type == "Debit":
                    category_totals[category]["debit"] += amount
                elif transaction_type == "Credit":
                    category_totals[category]["credit"] += amount
            else:
                category_totals[category] = {"debit": 0, "credit": 0}
                if transaction_type == "Debit":
                    category_totals[category]["debit"] = amount
                elif transaction_type == "Credit":
                    category_totals[category]["credit"] = amount

    filtered_category_totals = {k: v for k, v in category_totals.items() if v["debit"] != 0 or v["credit"] != 0}

    if filtered_category_totals:
        labels = list(filtered_category_totals.keys())
        net_totals = [v["debit"] - v["credit"] for v in filtered_category_totals.values()]

        greater_info = []
        for label, total in filtered_category_totals.items():
            if total["debit"] > total["credit"]:
                info = f"{label.capitalize()} (Debit is greater)"
            elif total["debit"] < total["credit"]:
                info = f"{label.capitalize()} (Credit is greater)"
            else:
                info = f"{label.capitalize()} (Equal Debit and Credit)"
            greater_info.append(info)

        abs_net_totals = np.abs(net_totals)

        plt.figure(figsize=(10, 10))
        ax1 = plt.subplot(121)

        wedges, texts, autotexts = ax1.pie(
            abs_net_totals,
            labels=None,
            autopct='%1.1f%%',
            startangle=140,
            textprops={'color': "w"},
            pctdistance=0.85,
        )
        plt.subplots_adjust(left=0.2, bottom=0.3, right=1, top=0.7, wspace=0)

        ax1.axis('equal')

        ax2 = plt.subplot(122)
        ax2.axis('off')

        legend_data = [(text, wedge.get_facecolor()) for text, wedge in zip(greater_info, wedges)]
        ax2.legend(handles=[plt.Line2D([0], [0], marker='o', color='w', label=label, markerfacecolor=color)
                            for label, color in legend_data], loc='center', title='Category Legend')

        plt.show()
    else:
        messagebox.showinfo("Info", "No category expenses found.")

pie_chart_button = ttk.Button(app, text="Create Category Pie Chart", command=create_category_pie_chart, style="OrangeHover.TButton")
pie_chart_button.pack(pady=10)

update_totals()

app.mainloop()
